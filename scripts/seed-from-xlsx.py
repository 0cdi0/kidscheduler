#!/usr/bin/env python3
"""
One-off conversion of the original 'Jahresplanung 2026/27 Betreuungskalender.xlsx'
(Daten + Legende sheets) into server/data/schedule.json.

Run once with: python3 scripts/seed-from-xlsx.py <path-to-xlsx>
Not needed at app runtime - the app reads/writes schedule.json directly from then on.
"""
import json
import re
import sys
from datetime import date

import openpyxl

KIDS = [
    {"id": "philipp", "name": "Philipp", "group": "dad", "color": "#6B4C9A"},
    {"id": "johannes", "name": "Johannes", "group": "dad", "color": "#B9A3D9"},
    {"id": "aleks", "name": "Aleks", "group": "mom", "color": "#2E6F9E"},
    {"id": "luis", "name": "Luis", "group": "mom", "color": "#8FC6E8"},
]

GROUPS = [
    {"id": "dad", "label": "Philipp & Johannes"},
    {"id": "mom", "label": "Aleks & Luis"},
]


def parse_de_date(s):
    d, m, y = s.split(".")
    return date(int(y), int(m), int(d)).isoformat()


def main(xlsx_path, out_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Birthdays from Legende!B40:D48
    legende = wb["Legende"]
    birthdays = {}
    for r in range(40, 49):
        d_raw = legende.cell(row=r, column=2).value
        name_raw = legende.cell(row=r, column=3).value
        note = legende.cell(row=r, column=4).value
        if not d_raw or not name_raw:
            continue
        iso = parse_de_date(str(d_raw).strip())
        name = re.sub(r"^[^\w]*", "", str(name_raw)).strip()  # strip leading cake emoji
        birthdays.setdefault(iso, []).append(
            {"name": name, "note": note or None}
        )

    ws = wb["Daten"]
    days = {}
    for row in ws.iter_rows(min_row=2, max_row=367, min_col=1, max_col=12):
        dt, weekday, kw, month, al, phj, status, holiday, break_, geb, notes, unsicher = [
            c.value for c in row
        ]
        if dt is None:
            continue
        iso = dt.date().isoformat()

        uncertain = bool(status) and "(?)" in str(status)

        kid_status = {}
        if al == "Ja":
            kid_status["aleks"] = "uncertain" if uncertain else "with-us"
            kid_status["luis"] = "uncertain" if uncertain else "with-us"
        if phj == "Ja":
            kid_status["philipp"] = "uncertain" if uncertain else "with-us"
            kid_status["johannes"] = "uncertain" if uncertain else "with-us"

        days[iso] = {
            "kids": kid_status,
            "holiday": holiday or None,
            "schoolBreak": break_ or None,
            "birthdays": birthdays.get(iso, []),
            "notes": notes or None,
        }

    schedule = {
        "meta": {
            "schoolYear": "2026/27",
            "source": "Jahresplanung_202627_Betreuungskalender.xlsx",
            "start": "2026-08-31",
            "end": "2027-08-31",
        },
        "kids": KIDS,
        "groups": GROUPS,
        "days": days,
    }

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(schedule, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"Wrote {len(days)} days to {out_path}")


if __name__ == "__main__":
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "Jahresplanung_202627_Betreuungskalender.xlsx"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "server/data/schedule.json"
    main(xlsx_path, out_path)
